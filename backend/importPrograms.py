#!/usr/bin/env python3
"""
UNIROUTE - importPrograms.py

Imports verified university program listings from an .xlsx workbook into
university_programs.

    python importPrograms.py <file.xlsx> [--dry-run]

Expects a "Programs" sheet with these headers (order-independent):
    University ID | University | Country | Degree Level | Program Name
    Official Source | Verification Status

Safety
------
* A row is only imported if its University ID exists AND the university name
  in the sheet matches the one in the database. Universities get renumbered
  whenever seed.js re-runs, so an ID alone is not enough to trust.
* Idempotent: rows upsert on (university_id, name, degree_level), so
  re-importing a corrected sheet updates rather than duplicates.
* --dry-run reports exactly what would happen and writes nothing.

Python deps: openpyxl, psycopg2 (both already present).
"""

import os
import sys
import unicodedata

import openpyxl
import psycopg2
from psycopg2.extras import execute_batch

REQUIRED = ["University ID", "University", "Degree Level", "Program Name"]

DEGREE = {
    "bachelor's": "Bachelor's", "bachelor": "Bachelor's", "bachelors": "Bachelor's",
    "master's": "Master's", "master": "Master's", "masters": "Master's",
    "phd": "PhD", "doctorate": "PhD", "doctoral": "PhD",
    "n/a": "N/A", "": "N/A",
}


def load_env(path):
    """Minimal .env reader so this matches what the Node backend connects to."""
    env = {}
    if not os.path.exists(path):
        return env
    for line in open(path, encoding="utf-8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def norm(s):
    """Casefold + strip accents, so 'Fundação' matches 'Fundacao'."""
    s = str(s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def norm_degree(v):
    return DEGREE.get(str(v or "").strip().lower(), str(v or "").strip() or "N/A")


def read_sheet(path, sheet="Programs"):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"No '{sheet}' sheet in {path}. Found: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Sheet is empty")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = [c for c in REQUIRED if c not in header]
    if missing:
        sys.exit(f"Missing required column(s): {missing}\nFound: {header}")
    idx = {name: i for i, name in enumerate(header)}
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({name: (r[i] if i < len(r) else None) for name, i in idx.items()})
    return out


def make_field_matcher(fields):
    """
    Conservative categorisation: only assign a field when the program name
    actually contains that field's name. Anything ambiguous stays NULL --
    a wrong field would list a university under a subject it doesn't teach,
    which is worse than no category at all.
    """
    ranked = sorted(
        ({"id": f_id, "key": norm(f_name)} for f_id, f_name in fields),
        key=lambda f: len(f["key"]),
        reverse=True,  # longest first: "computer science" wins over "science"
    )

    def match(program_name):
        p = norm(program_name)
        if not p:
            return None
        for f in ranked:
            if len(f["key"]) >= 4 and f["key"] in p:
                return f["id"]
        return None

    return match


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry_run = "--dry-run" in sys.argv
    if not args:
        sys.exit("Usage: python importPrograms.py <file.xlsx> [--dry-run]")
    path = args[0]
    if not os.path.exists(path):
        sys.exit(f"File not found: {path}")

    rows = read_sheet(path)
    print(f"Read {len(rows)} program rows from {os.path.basename(path)}")

    env = load_env(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
    conn = psycopg2.connect(
        host=env.get("DB_HOST", "127.0.0.1"), port=env.get("DB_PORT", 5432),
        dbname=env.get("DB_NAME"), user=env.get("DB_USER"), password=env.get("DB_PASSWORD"),
    )
    cur = conn.cursor()

    cur.execute("SELECT id, name FROM fields")
    match_field = make_field_matcher(cur.fetchall())

    uni_ids = sorted({int(r["University ID"]) for r in rows
                      if str(r["University ID"] or "").strip().isdigit()})
    cur.execute("SELECT id, name FROM universities WHERE id = ANY(%s)", (uni_ids,))
    db_unis = {i: n for i, n in cur.fetchall()}

    to_insert, problems = [], []
    mapped = unmapped = 0

    for r in rows:
        raw_id = str(r["University ID"] or "").strip()
        name = str(r["Program Name"] or "").strip()
        if not raw_id.isdigit() or not name:
            problems.append(f"missing id/name: {str(r)[:70]}")
            continue
        uid = int(raw_id)

        db_name = db_unis.get(uid)
        if db_name is None:
            problems.append(f"id {uid}: not present in universities table")
            continue
        if norm(db_name) != norm(r["University"]):
            problems.append(f'id {uid}: name mismatch - sheet "{r["University"]}" vs db "{db_name}"')
            continue

        fid = match_field(name)
        mapped += fid is not None
        unmapped += fid is None
        to_insert.append((
            uid, name, norm_degree(r["Degree Level"]), fid,
            r.get("Official Source") or None,
            r.get("Verification Status") or None,
            os.path.basename(path),
        ))

    print(f"\n  importable   : {len(to_insert)}")
    print(f"  skipped      : {len(problems)}")
    print(f"  field-mapped : {mapped}   (uncategorised: {unmapped} - imported, just no field)")
    if problems:
        print("\n  issues:")
        for p in problems[:12]:
            print("   -", p)
        if len(problems) > 12:
            print(f"   ... and {len(problems)-12} more")

    if dry_run:
        print("\n--dry-run: nothing written.")
        conn.rollback()
        return

    execute_batch(cur, """
        INSERT INTO university_programs
          (university_id, name, degree_level, field_id, url, verification, imported_from)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (university_id, name, degree_level) DO UPDATE
          SET field_id      = EXCLUDED.field_id,
              url           = EXCLUDED.url,
              verification  = EXCLUDED.verification,
              imported_from = EXCLUDED.imported_from
    """, to_insert, page_size=200)
    conn.commit()

    cur.execute("SELECT COUNT(*), COUNT(DISTINCT university_id) FROM university_programs")
    total, unis = cur.fetchone()
    print(f"\nWrote {len(to_insert)} rows.")
    print(f"university_programs now holds {total} programs across {unis} universities.")
    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
